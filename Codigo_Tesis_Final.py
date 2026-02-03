import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

# --- CONFIGURACIÓN DE LA PÁGINA ---
st.set_page_config(page_title="PSO Propagación - ULA", layout="wide")
st.title("🚀 Herramienta Computacional para la Aplicación de la Técnica PSO, " \
"en el Ajuste de Modelos de Propagación en Sistemas Inalámbricos de Comunicaciones ")

# --- FUNCIONES DE SOPORTE ---
def leer_rango_st(rango, hoja_excel):
    try:
        min_col, min_row, max_col, max_row = range_boundaries(rango)
        valores = []
        for row in hoja_excel.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.value is not None: valores.append(float(cell.value))
        return np.array(valores)
    except Exception as e:
        st.error(f"Error al leer rango: {e}"); return None

# --- FUNCIONES DE COSTO ---
def funcion_costo_lee(C, coef_ids, L_med, d, Lo, gamma, FA, do):
    c1, c2, c3 = 1.0, 1.0, 1.0
    for i, cid in enumerate(coef_ids):
        if cid == 1: c1 = C[i]
        elif cid == 2: c2 = C[i]
        elif cid == 3: c3 = C[i]
    L_est = (c1 * Lo) + (c2 * 10 * gamma * np.log10(d / do)) - (c3 * 10 * np.log10(FA))
    err = L_med - L_est
    return np.sqrt(np.sum(err**2) / (len(L_med) - 1))

def funcion_costo_hata(C, coef_ids, L_med, d, A, B, C_extra):
    c1, c2 = 1.0, 1.0
    for i, cid in enumerate(coef_ids):
        if cid == 1: c1 = C[i]
        elif cid == 2: c2 = C[i]
    L_est = c1 * 69.55 + A + (c2 * 44.90 + B) * np.log10(d) + C_extra
    err = L_med - L_est
    return np.sqrt(np.sum(err**2) / (len(L_med) - 1))
 
def funcion_costo_hata_extendido(C, coef_ids, L_med, d, A1, B1, C_extra):
    c1, c2 = 1.0, 1.0
    for i, cid in enumerate (coef_ids):
        if cid == 1: c1 = C[i]
        elif cid == 2: c2 = C[i]
    L_est = c1 * 46.30 + A1 + (c2 *44.90 + B1) * np.log10(d) + C_extra
    err = L_med - L_est
    return np.sqrt(np.sum(err**2) / (len(L_med) - 1))
 
# --- BARRA LATERAL ---
st.sidebar.header("📂 Carga de Datos")
archivo_subido = st.sidebar.file_uploader("Excel de mediciones", type=["xlsx"])

if archivo_subido:
    wb = load_workbook(archivo_subido, data_only=True)
    nombre_hoja = st.sidebar.selectbox("Hoja", wb.sheetnames)
    hoja = wb[nombre_hoja]
    
    c_d = st.sidebar.text_input("Rango Distancia (ej: A2:A23)", "A2:A23")
    c_L = st.sidebar.text_input("Rango Pérdida (ej: B2:B23)", "B2:B23")
    
    d_data = leer_rango_st(c_d, hoja)
    L_medido = leer_rango_st(c_L, hoja)

    st.sidebar.header("📡 Parámetros del Sistema")
    f = st.sidebar.number_input("Frecuencia (MHz)", value=600.0)
    hT = st.sidebar.number_input("Altura Antena Transmisora hT (m)", value=30.0)
    hR = st.sidebar.number_input("Altura Antena Receptora  hR (m)", value=1.5)
    
    modelo = st.sidebar.selectbox("Modelo", ["Lee", "Okumura-Hata", "Hata Extendido"])
    
    with st.expander("🛠️ Configuración del Modelo", expanded=True):
        if modelo == "Lee":
            c1_col, c2_col, c3_col = st.columns(3)
            Lo_orig = c1_col.number_input("Lo (dB)", value=125.0)
            gamma_orig = c2_col.number_input("Gamma (γ)", value=3.41)
            do = c3_col.number_input("do (km)", value=1.0)
            
            g_cols = st.columns(2)
            GT_dBi = g_cols[0].number_input("GT (dBi)", value=0.0)
            GR_dBi = g_cols[1].number_input("GR (dBi)", value=0.0)
            
            # Cálculo FA 
            FhT = (hT / 30.48)**2
            FhR = (hR / 3)**(2 if hR > 3 else 1)
            Ff = (f / 900)**(-2 if f <= 450 else -3)
            FGT = (10**(GT_dBi/10)) / 4
            FGR = 10**(GR_dBi/10)
            FA_orig = FhT * FhR * Ff * FGT * FGR
            
            c_opts = st.multiselect("Optimizar:", [1, 2, 3], default=[1, 2, 3], format_func=lambda x: f"c{x}")
        else:
            tipo_area = st.selectbox("Área", [1, 2, 3, 4], format_func=lambda x: {1:"Urbana Grande", 2:"Urbana Mediana/Pequeña", 3:"Suburbana", 4:"Rural"}[x])
            A_h = (1.1*np.log10(f)-0.7)*hR - (1.56*np.log10(f)-0.8)
            if modelo == "Okumura-Hata":
                A = 26.16*np.log10(f) - 13.82*np.log10(hT) - A_h
                B = 6.55*np.log10(hT)
            else: # Extendido
                Cm = 3 if (tipo_area == 1 and f > 1500) else 0
                A1 = 33.9*np.log10(f) - 13.82*np.log10(hT) - A_h + Cm
                B1 = 6.55*np.log10(hT)
            
            Cs = -2*(np.log10(f/28))**2 - 5.4
            Cr = -4.78*(np.log10(f))**2 + 18.33*np.log10(f) - 40.94
            C_extra = Cs if tipo_area == 3 else (Cr if tipo_area == 4 else 0)
            c_opts = st.multiselect("Optimizar:", [1, 2], default=[1, 2], format_func=lambda x: f"c{x}")

    # --- PARÁMETROS PSO CONFIG ---
    st.sidebar.header("⚙️ PARÁMETROS PSO ")
    n_part = st.sidebar.number_input("Partículas", value=40)
    n_ite = st.sidebar.number_input("Iteraciones", value=300)
    c1_pso = st.sidebar.number_input("Coeficiente Cognitivo", value = 2.0) 
    c2_pso = st.sidebar.number_input("Coeficiente Social", value = 2.0)
    w_max = st.sidebar.number_input("Ingrese el valor de Inercia Máxima W_max", value = 0.9)
    w_min =st.sidebar.number_input("Ingrese el valor de Inercia Mínima W_min", value = 0.4) 

    if st.button("🚀 Ejecutar Ajuste"):
        if d_data is not None and L_medido is not None:
            n_var = len(c_opts)
            
            # --- LÓGICA DE LÍMITES  ---
            limites = np.zeros((n_var, 2))
            for i, cid in enumerate(c_opts):
                if modelo == "Lee":
                    if cid == 1: limites[i] = [1.27, 1.29]
                    elif cid == 3: limites[i] = [0.84, 0.89]
                    else: limites[i] = [0.1, 5.0] # Para gamma (c2)
                else:
                    limites[i] = [0.01, 3.5]

            # Inicialización
            posicion = np.zeros((n_var, n_part))
            for i in range(n_var):
                posicion[i,:] = np.random.uniform(limites[i,0], limites[i,1], n_part)
            
            velocidad = np.zeros((n_var, n_part))
            mejor_pos_local = posicion.copy()
            costo_local = np.array([float('inf')]*n_part)

            def evaluar(p):
                if modelo == "Lee":
                    return funcion_costo_lee(p, c_opts, L_medido, d_data, Lo_orig, gamma_orig, FA_orig, do)
                elif modelo == "Okumura-Hata":
                    return funcion_costo_hata(p, c_opts, L_medido, d_data, A, B, C_extra)
                else:
                    return funcion_costo_hata_extendido(p, c_opts, L_medido, d_data, A1, B1, C_extra)

            mejor_costo_global = float('inf')
            mejor_pos_global = None
            historial = []

            # Bucle PSO
            bar = st.progress(0)
            for ite in range(n_ite):
                for i in range(n_part):
                    c_act = evaluar(posicion[:, i])
                    if c_act < costo_local[i]:
                        costo_local[i] = c_act
                        mejor_pos_local[:, i] = posicion[:, i]
                    if c_act < mejor_costo_global:
                        mejor_costo_global = c_act
                        mejor_pos_global = posicion[:, i].copy()

                w = w_max - (w_max - w_min) * (ite / n_ite)
                r1, r2 = np.random.rand(n_var, n_part), np.random.rand(n_var, n_part)
                velocidad = (w * velocidad + c1_pso * r1 * (mejor_pos_local - posicion) + 
                             c2_pso * r2 * (mejor_pos_global.reshape(-1,1) - posicion))
                posicion += velocidad

                # Clamping
                if n_var == 3:
                    for i in range(n_var):           
                        posicion[i, posicion[i, :] < limites[i, 0]] = limites[i, 0]
                        posicion[i, posicion[i, :] > limites[i, 1]] = limites[i, 1]

                posicion[posicion <= 0] = 1e-6
                historial.append(mejor_costo_global)
                bar.progress((ite + 1) / n_ite)
            
            # --- 1. Calcular el RMSE Inicial (Sin Optimizar) ---
            # Se creó un vector de coeficientes "neutros" (todos 1.0)
            c_iniciales = [1.0, 1.0, 1.0]
            c_ids_todos = [1, 2, 3]

            if modelo == "Lee":
                rmse_inicial = funcion_costo_lee(c_iniciales, c_ids_todos, L_medido, d_data, Lo_orig, gamma_orig, FA_orig, do)
            elif modelo == "Okumura-Hata":
            # Para Hata solo se usan 2 coeficientes
                rmse_inicial = funcion_costo_hata([1.0, 1.0], [1, 2], L_medido, d_data, A, B, C_extra)
            else: # Hata Extendido
                rmse_inicial = funcion_costo_hata_extendido([1.0, 1.0], [1, 2], L_medido, d_data, A1, B1, C_extra)

            # --- 2. Calcular la Mejora Porcentual ---
            mejora_pct = ((rmse_inicial - mejor_costo_global) / rmse_inicial) * 100

            # --- 3. Mostrar la Evidencia de mejora del modelo ----
            st.header("📊 Evidencia de Mejora del Modelo")

            col1, col2, col3 = st.columns(3)
            col1.metric("RMSE Inicial (Estándar)", f"{rmse_inicial:.4f} dB")
            col2.metric("RMSE Final (Optimizado)", f"{mejor_costo_global:.4f} dB", f"-{mejora_pct:.2f}%", delta_color="normal")
            col3.metric("Mejora en Precisión", f"{mejora_pct:.2f}%")

            # --- 4. Tabla Comparativa de Coeficientes ---
            datos_comparativa = {
            "Parámetro": [f"c{c}" for c in c_opts],
            "Valor Original (Teórico)": [1.0] * len(c_opts),
            "Valor Optimizado (PSO)": mejor_pos_global
            }
            st.table(pd.DataFrame(datos_comparativa))

            st.info(f"💡** El uso del algoritmo PSO permitió reducir la incertidumbre del modelo en un {mejora_pct:.2f}%, "
                f"evidenciando que los coeficientes c1, c2 y c3 efectivamente sintonizan la ecuación teórica con la realidad del terreno.")

            st.success(f"Ajuste completado - RMSE Final: {mejor_costo_global:.4f} dB")
            
            c1f = mejor_pos_global[0] if 1 in c_opts else 1.0
            c2f = mejor_pos_global[1] if 2 in c_opts else 1.0 # Ajusta el índice según tu lógica

            st.subheader("📈 Validación: Mediciones vs. Modelo")
            fig_prop, ax_prop = plt.subplots(figsize=(10, 5))
            
            # Datos reales (los 2,000+ puntos)
            ax_prop.scatter(d_data, L_medido, color='red', s=8, alpha=0.3, label='Datos Experimentales')
            
            # Línea del modelo
            d_linea = np.linspace(min(d_data), max(d_data), 100)
            if modelo == "Okumura-Hata":
                L_linea = c1f * 69.55 + A + (c2f * 44.90 + B) * np.log10(d_linea) + C_extra
            elif modelo == "Hata Extendido":
                L_linea = c1f * 46.30 + A1 + (c2f * 44.90 + B1) * np.log10(d_linea) + C_extra
            else: # Lee
                c3f = mejor_pos_global[2] if 3 in c_opts else 1.0
                L_linea = (c1f * Lo_orig) + (c2f * 10 * gamma_orig * np.log10(d_linea / do)) - (c3f * 10 * np.log10(FA_orig))

            ax_prop.set_xlabel("Distancia (km)")
            ax_prop.set_ylabel("Pérdida de Propagación (dB)")
            ax_prop.legend()
            ax_prop.grid(True, alpha=0.3)
            st.pyplot(fig_prop)

            # 3. GRÁFICA DE CONVERGENCIA 
            st.subheader("📉 Análisis de Convergencia del Algoritmo")
            fig_conv, ax_conv = plt.subplots(figsize=(10, 4))
            ax_conv.plot(historial, color='green')
            ax_conv.set_title("Evolución del RMSE durante la Optimización")
            ax_conv.set_xlabel("Iteración")
            ax_conv.set_ylabel("RMSE (dB)")
            st.pyplot(fig_conv)

            # Resultados
            st.success(f"Ajuste completado - RMSE Final: {mejor_costo_global:.4f} dB")
            c1_res, c2_res = st.columns(2)
            with c1_res:
                st.write("**Coeficientes Optimizados:**")
                st.table(pd.DataFrame({"Parámetro": [f"c{c}" for c in c_opts], "Valor": mejor_pos_global}))
            
            